#!/usr/bin/env python3
"""Build the wholesale scaling what-if model."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

F = "Arial"
BLUE = Font(name=F, size=10, color="0000FF")            # hardcoded input
BLACK = Font(name=F, size=10)                           # formula
GREEN = Font(name=F, size=10, color="008000")           # link to another sheet
BOLD = Font(name=F, size=10, bold=True)
TITLE = Font(name=F, size=14, bold=True)
HEAD = Font(name=F, size=10, bold=True, color="FFFFFF")
SMALL = Font(name=F, size=8, italic=True, color="595959")
YFILL = PatternFill("solid", fgColor="FFFF00")          # assumption to review
HFILL = PatternFill("solid", fgColor="1F3864")
SFILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))
USD = '$#,##0.00;($#,##0.00);-'
USD0 = '$#,##0;($#,##0);-'
PCT = '0.0%'
NUM2 = '#,##0.00'
NUM1 = '#,##0.1'

wb = openpyxl.Workbook()


def sheet(name, widths):
    ws = wb.create_sheet(name) if name not in wb.sheetnames else wb[name]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    return ws


def title(ws, text, sub=None, span=8):
    ws["A1"] = text
    ws["A1"].font = TITLE
    if sub:
        ws["A2"] = sub
        ws["A2"].font = SMALL


def hdr(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.font = HEAD
        c.fill = HFILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 28


def band(ws, row, text, span=9):
    c = ws.cell(row=row, column=1, value=text)
    c.font = BOLD
    c.fill = SFILL
    for i in range(2, span + 1):
        ws.cell(row=row, column=i).fill = SFILL


def put(ws, ref, val, font=BLACK, fmt=None, fill=None, note=None):
    c = ws[ref]
    c.value = val
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if note:
        c.comment = Comment(note, "model")
    return c


# ============================================================ INPUTS
ws = sheet("Inputs", [42, 14, 13, 13, 60])
wb.remove(wb["Sheet"])
title(ws, "WHOLESALE SCALING MODEL — Here & There",
      "Every blue cell is yours to change. Yellow = assumption with no source data behind it. "
      "Black = formula, don't type over it.")

put(ws, "A4", "HOW TO USE", BOLD)
for i, t in enumerate([
    "1. Set the three volume scenarios in B7:B9, then pick the active one in B11 (dropdown).",
    "2. Set wholesale % of MSRP in B14. MSRPs live on the Items tab.",
    "3. Choose the cost basis in B18: 'Flat' uses your $1.00; 'Computed' uses the real recipe "
    "+ labor + overhead stack from the CostStack tab.",
    "4. Read Scaling for the 1→5 item build, Scenarios for bear/base/bull side by side.",
]):
    put(ws, f"A{5+i}", t, SMALL)

band(ws, 10, "VOLUME SCENARIOS  (units sold per day, per item)", 5)
put(ws, "A7", "Bear — units/day per item", BOLD)
put(ws, "A8", "Base — units/day per item", BOLD)
put(ws, "A9", "Bull — units/day per item", BOLD)
put(ws, "B7", 7, BLUE, NUM2)
put(ws, "B8", 10, BLUE, NUM2)
put(ws, "B9", 15, BLUE, NUM2)
put(ws, "C7", "your bear case", SMALL)
put(ws, "C8", "your base case", SMALL)
put(ws, "C9", "your bull case", SMALL)

put(ws, "A11", "ACTIVE SCENARIO", BOLD)
put(ws, "B11", "Base", BLUE, fill=YFILL)
put(ws, "A12", "Units/day per item (active)", BOLD)
put(ws, "B12", '=IF($B$11="Bear",$B$7,IF($B$11="Bull",$B$9,$B$8))', BLACK, NUM2)
dv = DataValidation(type="list", formula1='"Bear,Base,Bull"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(ws["B11"])

band(ws, 13, "PRICING", 5)
put(ws, "A14", "Wholesale price as % of MSRP", BOLD)
put(ws, "B14", 0.50, BLUE, PCT)
put(ws, "C14", "your instruction: wholesale at 50%", SMALL)

band(ws, 16, "COST BASIS", 5)
put(ws, "A17", "Flat cost per unit (your $1 assumption)", BOLD)
put(ws, "B17", 1.00, BLUE, USD)
put(ws, "A18", "Cost basis", BOLD)
put(ws, "B18", "Flat", BLUE, fill=YFILL)
dv2 = DataValidation(type="list", formula1='"Flat,Computed"', allow_blank=False)
ws.add_data_validation(dv2)
dv2.add(ws["B18"])
put(ws, "C18", "Flat = your $1.00 · Computed = real recipe + consumables", SMALL)
put(ws, "A19", "Computed materials + consumables / unit", BOLD)
put(ws, "B19", "=CostStack!$D$83", GREEN, USD)
put(ws, "A20", "Cost per unit in use", BOLD)
put(ws, "B20", '=IF($B$18="Computed",$B$19,$B$17)', BLACK, USD)
put(ws, "A21", "Include labor + kitchen time cost?", BOLD)
put(ws, "B21", "Yes", BLUE, fill=YFILL)
dv3 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
ws.add_data_validation(dv3)
dv3.add(ws["B21"])
put(ws, "C21", "No = you absorb your own time and already pay the kitchen anyway", SMALL)
put(ws, "D19", "Excludes labor and kitchen time — those are costed by the hour in Scaling, so "
    "there is no double-count either way you set B18.", SMALL)

put(ws, "A22", "Wholesale accounts (stores)", BOLD)
put(ws, "B22", 1, BLUE, NUM2, YFILL)
put(ws, "C22", "Here & There = 1 today", SMALL)
put(ws, "D22", "Units/day figures above are PER ITEM, PER ACCOUNT. More accounts fill the same "
    "batches; more SKUs at one account do not.", SMALL)
band(ws, 23, "LABOR AND BAKE CONSTANTS  (from the Kairos app — real, not assumed)", 5)
put(ws, "A24", "Batch yield (units per batch)", BOLD)
put(ws, "B24", 20, BLUE, NUM2)
put(ws, "C24", "BATCH_YIELD, data-entry-app/index.html", SMALL)
put(ws, "A25", "Setup (min, once per bake session)", BOLD)
put(ws, "B25", 20, BLUE, NUM2)
put(ws, "C25", "SETUP_MIN", SMALL)
put(ws, "A26", "Prep per batch (min)", BOLD)
put(ws, "B26", 15, BLUE, NUM2)
put(ws, "C26", "PREP_MIN", SMALL)
put(ws, "A27", "Bake per batch (min)", BOLD)
put(ws, "B27", 14, BLUE, NUM2)
put(ws, "C27", "BAKE_MIN — single oven, glaze overlaps the bake", SMALL)
put(ws, "A28", "Labor wage ($/hour, loaded)", BOLD)
put(ws, "B28", 30.00, BLUE, USD)
put(ws, "C28", "your figure", SMALL)
put(ws, "A29", "Commercial kitchen ($/hour)", BOLD)
put(ws, "B29", 50.00, BLUE, USD)
put(ws, "C29", "your figure — charged on bake time, not per unit", SMALL)
put(ws, "A30", "Combined time cost ($/hour)", BOLD)
put(ws, "B30", "=$B$28+$B$29", BLACK, USD)
put(ws, "A31", "Setup treatment", BOLD)
put(ws, "B31", "Shared", BLUE, fill=YFILL)
dv4 = DataValidation(type="list", formula1='"Per item,Shared"', allow_blank=False)
ws.add_data_validation(dv4)
dv4.add(ws["B31"])
put(ws, "C31", "Per item = each SKU gets its own session. Shared = all SKUs in one session.", SMALL)
put(ws, "A32", "Bake sessions per week", BOLD)
put(ws, "B32", 3, BLUE, NUM2, YFILL)
put(ws, "D32", "THE BIGGEST LEVER. Baking 7x/week means tiny half-empty batches; 2-3x/week "
    "fills them and divides the setup across far more units.", SMALL)

band(ws, 33, "CALENDAR", 5)
put(ws, "A34", "Operating days per week", BOLD)
put(ws, "B34", 7, BLUE, NUM2)
put(ws, "A35", "Days per year", BOLD)
put(ws, "B35", "=$B$34*52", BLACK, NUM2)
put(ws, "A36", "Days per month", BOLD)
put(ws, "B36", "=$B$35/12", BLACK, NUM2)

band(ws, 38, "REALITY CHECKS  (measured, for context — these do not feed the model)", 5)
checks = [
    ("Madeleine retail price observed", 4.50, "median of 76 single-madeleine LM receipts", USD),
    ("App's stored madeleine price", 2.50, "push_store_stats.py PRICE — equals 50% of a $5 MSRP", USD),
    ("LM madeleines/day, Aug 1-7", 8.7, "Store Sales tab", NUM2),
    ("LM store transactions/day, Aug 1-7", 233, "Store Sales tab", NUM2),
]
for i, (l, v, s, f) in enumerate(checks):
    put(ws, f"A{39+i}", l, BLACK)
    put(ws, f"B{39+i}", v, BLUE, f)
    put(ws, f"C{39+i}", s, SMALL)

# ============================================================ ITEMS
ws = sheet("Items", [10, 26, 12, 14, 16, 16, 14, 12, 16, 16])
title(ws, "ITEMS", "MSRPs are yours. Wholesale = MSRP x Inputs!B14. Override a single item's "
                   "cost in column E; leave blank to use the global basis.")
hdr(ws, 4, ["#", "Item", "MSRP", "Wholesale $", "Cost override", "Cost used $/unit",
            "Margin $/unit", "Margin %", "Cum. wholesale $", "Cum. cost $"])
msrp = [5, 7, 6, 5, 5]
for i, m in enumerate(msrp):
    r = 5 + i
    put(ws, f"A{r}", i + 1, BLACK)
    put(ws, f"B{r}", f"Item {i+1}", BLUE)
    put(ws, f"C{r}", m, BLUE, USD)
    put(ws, f"D{r}", f"=$C{r}*Inputs!$B$14", BLACK, USD)
    put(ws, f"E{r}", None, BLUE, USD, YFILL)
    put(ws, f"F{r}", f"=IF(ISNUMBER($E{r}),$E{r},Inputs!$B$20)", BLACK, USD)
    put(ws, f"G{r}", f"=$D{r}-$F{r}", BLACK, USD)
    put(ws, f"H{r}", f"=IFERROR($G{r}/$D{r},0)", BLACK, PCT)
    put(ws, f"I{r}", f"=SUM($D$5:$D{r})", BLACK, USD)
    put(ws, f"J{r}", f"=SUM($F$5:$F{r})", BLACK, USD)
    for col in "ABCDEFGHIJ":
        ws[f"{col}{r}"].border = THIN
put(ws, "B11", "TOTAL / all 5", BOLD)
for col, f in (("C", "=SUM($C$5:$C$9)"), ("D", "=SUM($D$5:$D$9)"), ("F", "=SUM($F$5:$F$9)"),
               ("G", "=SUM($G$5:$G$9)")):
    put(ws, f"{col}11", f, BOLD, USD)
put(ws, "H11", "=IFERROR($G$11/$D$11,0)", BOLD, PCT)
put(ws, "A13", "MSRPs supplied by you: item 1 $5, item 2 $7, item 3 $6, item 4 $5, item 5 $5.", SMALL)
put(ws, "A14", "Rename the items in column B — the model does not care what they are called.", SMALL)

# ============================================================ SCALING
ws = sheet("Scaling", [26, 11, 13, 13, 12, 13, 13, 13, 13, 12, 13, 13])
title(ws, "SCALING 1 → 5 ITEMS",
      "Weekly basis, because you bake in sessions, not continuously. Revenue scales linearly with "
      "items; bake time does not. That gap is the whole story.")
put(ws, "A3", '=" Scenario: "&Inputs!$B$11&"  ·  "&TEXT(Inputs!$B$12,"0.0")&" units/day/item  ·  "'
              '&TEXT(Inputs!$B$32,"0")&" bake sessions/week  ·  setup "&Inputs!$B$31&"  ·  cost basis "'
              '&Inputs!$B$18', BOLD)

hdr(ws, 5, ["Items sold", "Units/week", "Units per session, per item", "Batches per session, per item",
            "Bake hrs/week", "Revenue/week", "Materials/week", "Labor+kitchen/week", "Profit/week",
            "Profit/day", "Profit/year", "Incremental /yr"])
UPS = "Inputs!$B$22*Inputs!$B$12*7/Inputs!$B$32"
BPS = f"ROUNDUP({UPS}/Inputs!$B$24,0)"
for n in range(1, 6):
    r = 5 + n
    put(ws, f"A{r}", f"{n} item" + ("s" if n > 1 else ""), BOLD)
    put(ws, f"B{r}", f"={n}*Inputs!$B$22*Inputs!$B$12*7", BLACK, NUM2)
    put(ws, f"C{r}", f"={UPS}", BLACK, NUM2)
    put(ws, f"D{r}", f"={BPS}", BLACK, NUM2)
    put(ws, f"E{r}", f'=IF(Inputs!$B$31="Shared",'
                     f'Inputs!$B$32*(Inputs!$B$25+{n}*{BPS}*(Inputs!$B$26+Inputs!$B$27))/60,'
                     f'Inputs!$B$32*{n}*(Inputs!$B$25+{BPS}*(Inputs!$B$26+Inputs!$B$27))/60)',
        BLACK, NUM2)
    put(ws, f"F{r}", f"=Inputs!$B$22*Inputs!$B$12*7*INDEX(Items!$I$5:$I$9,{n})", BLACK, USD)
    put(ws, f"G{r}", f"=Inputs!$B$22*Inputs!$B$12*7*INDEX(Items!$J$5:$J$9,{n})", BLACK, USD)
    put(ws, f"H{r}", f'=IF(Inputs!$B$21="Yes",$E{r}*Inputs!$B$30,0)', BLACK, USD)
    put(ws, f"I{r}", f"=$F{r}-$G{r}-$H{r}", BLACK, USD)
    put(ws, f"J{r}", f"=$I{r}/7", BLACK, USD)
    put(ws, f"K{r}", f"=$I{r}*52", BLACK, USD0)
    put(ws, f"L{r}", "-" if n == 1 else f"=$K{r}-$K{r-1}", BLACK, USD0)
    for col in "ABCDEFGHIJKL":
        ws[f"{col}{r}"].border = THIN

put(ws, "A12", "PER-UNIT AND PER-HOUR VIEW", BOLD)
hdr(ws, 13, ["Items sold", "Profit $/unit", "Profit $/bake hr", "Bake min/unit",
             "Time cost $/unit", "Materials $/unit", "Margin % of wholesale"])
for n in range(1, 6):
    r, s_ = 13 + n, 5 + n
    put(ws, f"A{r}", f"{n} item" + ("s" if n > 1 else ""), BOLD)
    put(ws, f"B{r}", f"=IFERROR($I{s_}/$B{s_},0)", BLACK, USD)
    put(ws, f"C{r}", f"=IFERROR($I{s_}/$E{s_},0)", BLACK, USD)
    put(ws, f"D{r}", f"=IFERROR($E{s_}*60/$B{s_},0)", BLACK, NUM2)
    put(ws, f"E{r}", f"=IFERROR($H{s_}/$B{s_},0)", BLACK, USD)
    put(ws, f"F{r}", f"=IFERROR($G{s_}/$B{s_},0)", BLACK, USD)
    put(ws, f"G{r}", f"=IFERROR($I{s_}/$F{s_},0)", BLACK, PCT)
    for col in "ABCDEFG":
        ws[f"{col}{r}"].border = THIN

put(ws, "A20", "THE THREE LEVERS, IN ORDER OF POWER", BOLD)
put(ws, "A21", "1. Bake sessions per week (Inputs!B32). Fewer, bigger sessions fill the 20-unit "
               "batches and pay the 20-min setup once instead of seven times.", SMALL)
put(ws, "A22", '2. Setup treatment (Inputs!B31). "Shared" bakes all five SKUs in one session — '
               "one setup for the lot, not five.", SMALL)
put(ws, "A23", "3. Batch yield (Inputs!B24). Bigger trays divide the 29 min of prep+bake across "
               "more units. See the BreakEven tab, section D.", SMALL)
put(ws, "A24", "Adding SKUs raises revenue in a straight line. Whether it raises PROFIT depends "
               "entirely on whether the new SKU rides an existing session or forces a new one.", SMALL)

# ============================================================ SCENARIOS
ws = sheet("Scenarios", [24, 12, 13, 14, 13, 14, 13, 14, 14])
title(ws, "BEAR / BASE / BULL", "All three volume cases at once, on the same weekly basis. "
                                "Each block reads its own units/day from Inputs.")
row = 4
BLOCKS = [("BEAR", "$B$7", 4), ("BASE", "$B$8", 12), ("BULL", "$B$9", 20)]
for label, ref, _ in BLOCKS:
    band(ws, row, f"{label}  —  units/day per item = Inputs!{ref}", 9)
    hdr(ws, row + 1, ["Items sold", "Units/week", "Bake hrs/week", "Revenue/week",
                      "Materials/week", "Labor+kitchen/week", "Profit/week", "Profit/day",
                      "Profit/year"])
    ups = f"Inputs!$B$22*Inputs!{ref}*7/Inputs!$B$32"
    bps = f"ROUNDUP({ups}/Inputs!$B$24,0)"
    for n in range(1, 6):
        r = row + 1 + n
        put(ws, f"A{r}", f"{n} item" + ("s" if n > 1 else ""), BOLD)
        put(ws, f"B{r}", f"={n}*Inputs!$B$22*Inputs!{ref}*7", BLACK, NUM2)
        put(ws, f"C{r}", f'=IF(Inputs!$B$31="Shared",'
                         f'Inputs!$B$32*(Inputs!$B$25+{n}*{bps}*(Inputs!$B$26+Inputs!$B$27))/60,'
                         f'Inputs!$B$32*{n}*(Inputs!$B$25+{bps}*(Inputs!$B$26+Inputs!$B$27))/60)',
            BLACK, NUM2)
        put(ws, f"D{r}", f"=Inputs!$B$22*Inputs!{ref}*7*INDEX(Items!$I$5:$I$9,{n})", BLACK, USD)
        put(ws, f"E{r}", f"=Inputs!$B$22*Inputs!{ref}*7*INDEX(Items!$J$5:$J$9,{n})", BLACK, USD)
        put(ws, f"F{r}", f'=IF(Inputs!$B$21="Yes",$C{r}*Inputs!$B$30,0)', BLACK, USD)
        put(ws, f"G{r}", f"=$D{r}-$E{r}-$F{r}", BLACK, USD)
        put(ws, f"H{r}", f"=$G{r}/7", BLACK, USD)
        put(ws, f"I{r}", f"=$G{r}*52", BLACK, USD0)
        for col in "ABCDEFGHI":
            ws[f"{col}{r}"].border = THIN
    row += 8

band(ws, row, "SUMMARY — annual profit at 5 items", 9)
for i, (label, _, base_row) in enumerate(BLOCKS):
    put(ws, f"A{row+1+i}", label.title(), BOLD)
    put(ws, f"B{row+1+i}", f"=$I${base_row+6}", BLACK, USD0)
put(ws, f"C{row+1}", f"=IFERROR($B{row+1}/$B{row+2}-1,0)", BLACK, PCT)
put(ws, f"C{row+3}", f"=IFERROR($B{row+3}/$B{row+2}-1,0)", BLACK, PCT)
put(ws, f"D{row+1}", "vs base", SMALL)
put(ws, f"D{row+3}", "vs base", SMALL)

# ============================================================ COST STACK
ws = sheet("CostStack", [34, 13, 12, 14, 13, 46])
title(ws, "WHAT A MADELEINE ACTUALLY COSTS",
      "Recipes and ingredient prices are real, from kairos.html. Labor is derived from the app's "
      "bake constants. Everything else had NO source data and is an assumption — yellow.")

band(ws, 4, "A · INGREDIENT PRICES   (source: kairos.html DEFAULT_PRICES — Costco / Target / Amazon retail packs)", 6)
hdr(ws, 5, ["Ingredient", "Pack size", "Unit", "Pack price", "$ per g/ml", "Note"])
ing = [
    ("Cake Flour (AP proxy)", 9071.8, "g", 19.99, "Cake Flour is UNPRICED in source; AP flour used"),
    ("Baking Powder", 1814.4, "g", 14.99, "Costco Clabber Girl 4lb"),
    ("Salt", 850.5, "g", 3.99, "Costco Kirkland sea salt 30oz"),
    ("Eggs", 3400.0, "g", 8.29, "Costco Kirkland cage free 60ct"),
    ("Sugar", 11340.0, "g", 20.19, "Costco C&H 25lb"),
    ("Vanilla Extract", 473.0, "ml", 10.87, "Costco Kirkland 16oz"),
    ("Unsalted Butter", 907.2, "g", 18.49, "Kerrygold 8oz x4"),
    ("Powdered Sugar", 907.2, "g", 3.89, "glazes"),
    ("Lemon Juice", 2000.0, "ml", 9.09, "Lemon Poppy glaze"),
    ("Coconut Milk", 400.0, "ml", 2.39, "Ube glaze"),
    ("Ube Extract", 60.0, "ml", 6.95, "Ube"),
    ("Cream Cheese", 1360.8, "g", 10.49, "Dot Cake frosting"),
    ("Nonpareil Sprinkles", 998.0, "g", 16.27, "2.2lb bag tops ~100 Dot Cakes"),
    ("Coconut Flakes", 453.6, "g", 8.59, "1lb bag tops ~150 Ube"),
    ("Matcha Powder", 340.0, "g", 18.32, "Matcha"),
    ("Cocoa Powder", 700.0, "g", 14.99, "Double Choco"),
]
for i, (n, amt, u, p, note) in enumerate(ing):
    r = 6 + i
    put(ws, f"A{r}", n, BLACK)
    put(ws, f"B{r}", amt, BLUE, NUM2)
    put(ws, f"C{r}", u, BLACK)
    put(ws, f"D{r}", p, BLUE, USD)
    put(ws, f"E{r}", f"=IFERROR($D{r}/$B{r},0)", BLACK, '$#,##0.00000')
    put(ws, f"F{r}", note, SMALL)
    ws[f"A{r}"].border = THIN
IROW = {n.split(" (")[0]: 6 + i for i, (n, *_rest) in enumerate(ing)}

band(ws, 23, "B · BASE BATTER, one batch (yields 20 madeleines) — everything except the flour", 6)
hdr(ws, 24, ["Ingredient", "Grams / ml", "", "Cost", "", ""])
base = [("Baking Powder", 6), ("Salt", 2), ("Eggs", 200), ("Sugar", 140),
        ("Vanilla Extract", 5), ("Unsalted Butter", 200)]
for i, (n, gq) in enumerate(base):
    r = 25 + i
    put(ws, f"A{r}", n, BLACK)
    put(ws, f"B{r}", gq, BLUE, NUM2)
    put(ws, f"D{r}", f"=$B{r}*$E${IROW[n]}", BLACK, USD)
put(ws, "A31", "Base ex-flour, per batch", BOLD)
put(ws, "D31", "=SUM($D$25:$D$30)", BOLD, USD)
put(ws, "F25", "MAD_BASE in kairos.html. Yield 20 = BATCH_YIELD in the app.", SMALL)

band(ws, 33, "C · COST PER MADELEINE BY FLAVOR   (ingredients only — no labor, no energy, no packaging)", 6)
hdr(ws, 34, ["Flavor", "Flour g", "Flavoring $/batch", "Glaze $/batch", "Topping $/unit",
             "Ingredients $/unit"])
flav = [
    ("Lemon Poppy", 174, "0", f"=100*$E${IROW['Powdered Sugar']}+25*$E${IROW['Lemon Juice']}+4*$E${IROW['Unsalted Butter']}", "0"),
    ("Sea Salt / Classic", 174, "0", "0", "0"),
    ("Ube", 134, f"=2.5*$E${IROW['Ube Extract']}",
     f"=100*$E${IROW['Powdered Sugar']}+50*$E${IROW['Coconut Milk']}+1*$E${IROW['Ube Extract']}",
     f"=($B${IROW['Coconut Flakes']}/150)*$E${IROW['Coconut Flakes']}"),
    ("Dot Cake", 174, "0", "0",
     f"=(113/200)*$E${IROW['Unsalted Butter']}+(225/200)*$E${IROW['Cream Cheese']}"
     f"+(70/200)*$E${IROW['Powdered Sugar']}+($B${IROW['Nonpareil Sprinkles']}/100)*$E${IROW['Nonpareil Sprinkles']}"),
    ("Matcha", 156, f"=14*$E${IROW['Matcha Powder']}", "0", "0"),
    ("Double Choco", 146, f"=28*$E${IROW['Cocoa Powder']}", "0", "0"),
]
for i, (n, fl, fv, gl, tp) in enumerate(flav):
    r = 35 + i
    put(ws, f"A{r}", n, BLACK)
    put(ws, f"B{r}", fl, BLUE, NUM2)
    put(ws, f"C{r}", fv, BLACK, USD)
    put(ws, f"D{r}", gl, BLACK, USD)
    put(ws, f"E{r}", tp, BLACK, USD)
    put(ws, f"F{r}", f"=($D$31+$B{r}*$E${IROW['Cake Flour']}+$C{r}+$D{r})/Inputs!$B$24+$E{r}",
        BLACK, USD)
    ws[f"A{r}"].border = THIN
put(ws, "A41", "Average across flavors", BOLD)
put(ws, "F41", "=AVERAGE($F$35:$F$40)", BOLD, USD)
put(ws, "A42", "UNPRICED in the source data: Lemon Extract 4g, Lemon Poppy Seed 6g (Lemon Poppy); "
               "Ube Powder 20g (Ube). Add them here if you have prices:", SMALL)
put(ws, "A43", "Extra flavoring cost per unit, not yet priced", BLACK)
put(ws, "B43", 0.00, BLUE, USD, YFILL)

band(ws, 45, "D · TIME COST — labor + commercial kitchen   (minutes are real; the two rates are yours)", 6)
put(ws, "A46", '=" Formula: minutes = "&Inputs!$B$25&" setup + batches x ("&Inputs!$B$26&'
               '" prep + "&Inputs!$B$27&" bake).  Batch yield "&Inputs!$B$24&"."', SMALL)
hdr(ws, 47, ["Units baked in a session", "Batches", "Minutes", "Hours", "Labor+kitchen $/unit", "Min/unit"])
for i, u in enumerate([10, 20, 40, 50, 100, 200, 300]):
    r = 48 + i
    put(ws, f"A{r}", u, BLUE, NUM2)
    put(ws, f"B{r}", f"=ROUNDUP($A{r}/Inputs!$B$24,0)", BLACK, NUM2)
    put(ws, f"C{r}", f"=Inputs!$B$25+$B{r}*(Inputs!$B$26+Inputs!$B$27)", BLACK, NUM2)
    put(ws, f"D{r}", f"=$C{r}/60", BLACK, NUM2)
    put(ws, f"E{r}", f"=IFERROR($D{r}*Inputs!$B$30/$A{r},0)", BLACK, USD)
    put(ws, f"F{r}", f"=IFERROR($C{r}/$A{r},0)", BLACK, NUM2)
    ws[f"A{r}"].border = THIN
put(ws, "A56", "Time cost $/unit used in the stack below", BOLD)
put(ws, "B56", "=Inputs!$B$12", BLACK, NUM2)
put(ws, "C56", "units/day/item, active scenario", SMALL)
put(ws, "D56", "=(Inputs!$B$25+ROUNDUP($B$56/Inputs!$B$24,0)*(Inputs!$B$26+Inputs!$B$27))/60"
               "*Inputs!$B$30/$B$56", BOLD, USD)

band(ws, 58, "E · EVERYTHING THE RECIPE DATA DOES NOT INCLUDE   (all assumptions — set these)", 6)
put(ws, "A59", "The source costing has ingredients only. There is no electricity, gas, packaging, "
               "rent, or overhead anywhere in it. Fill these in:", SMALL)
hdr(ws, 60, ["Line", "Input", "Unit", "$ per madeleine", "", "Basis / how to set it"])
oh = [
    ("Oven power draw", 2.5, "kW", None, "Typical convection deck oven. Measure yours."),
    ("Electricity rate", 0.32, "$/kWh", None, "SCE commercial ballpark, Aug 2026. Check your bill."),
    ("Packaging per unit", 0.15, "$", None, "Sleeve / box / label / sticker. Wholesale often needs more."),
    ("Delivery per unit", 0.10, "$", None, "Fuel and driver time to the wholesale account."),
    ("Other overhead", 0.00, "$", None, "Kitchen rent is now charged hourly above — put only non-kitchen overhead here."),
    ("Waste / shrink", 0.05, "%", None, "Fraction baked but not sold. Uplifts every cost above."),
]
for i, (n, v, u, _x, note) in enumerate(oh):
    r = 61 + i
    put(ws, f"A{r}", n, BLACK)
    put(ws, f"B{r}", v, BLUE, PCT if u == "%" else NUM2, YFILL)
    put(ws, f"C{r}", u, BLACK)
    put(ws, f"F{r}", note, SMALL)
    ws[f"A{r}"].border = THIN
put(ws, "D61", "=$B$61*((Inputs!$B$25+ROUNDUP($B$56/Inputs!$B$24,0)*Inputs!$B$27)/60)*$B$62/$B$56",
    BLACK, USD)
put(ws, "E61", "oven on for setup + bake only", SMALL)
put(ws, "D63", "=$B$63", BLACK, USD)
put(ws, "D64", "=$B$64", BLACK, USD)
put(ws, "D65", "=$B$65", BLACK, USD)

band(ws, 68, "F · ALL-IN COST PER MADELEINE", 6)
hdr(ws, 69, ["Component", "", "", "$ per unit", "", "Source"])
stack = [
    ("Ingredients (recipe average)", "=$F$41+$B$43", "REAL — kairos.html recipes and prices"),
    ("Labor + commercial kitchen", '=IF(Inputs!$B$21="Yes",$D$56,0)', "REAL minutes x ($30 labor + $50 kitchen)"),
    ("Oven electricity", "=$D$61", "ASSUMPTION"),
    ("Packaging", "=$D$63", "ASSUMPTION"),
    ("Delivery", "=$D$64", "ASSUMPTION"),
    ("Overhead allocation", "=$D$65", "ASSUMPTION"),
]
for i, (n, f, src) in enumerate(stack):
    r = 70 + i
    put(ws, f"A{r}", n, BLACK)
    put(ws, f"D{r}", f, BLACK, USD)
    put(ws, f"F{r}", src, SMALL)
    ws[f"A{r}"].border = THIN
put(ws, "A76", "ALL-IN COST PER UNIT", BOLD)
put(ws, "D76", "=SUM($D$70:$D$75)*(1+$B$66)", BOLD, USD)
put(ws, "F76", "waste uplift applied to the whole stack", SMALL)
put(ws, "A77", "Your flat assumption", BLACK)
put(ws, "D77", "=Inputs!$B$17", GREEN, USD)
put(ws, "A78", "Difference (all-in vs flat)", BOLD)
put(ws, "D78", "=$D$76-$D$77", BOLD, USD)
put(ws, "A79", "Wholesale price at 50% of a $5 MSRP", BLACK)
put(ws, "D79", "=5*Inputs!$B$14", BLACK, USD)
put(ws, "A80", "Contribution per unit at that price", BOLD)
put(ws, "D80", "=$D$79-$D$76", BOLD, USD)
put(ws, "A81", "Gross margin % at that price", BOLD)
put(ws, "D81", "=IFERROR($D$80/$D$79,0)", BOLD, PCT)
band(ws, 82, "G · MATERIALS + CONSUMABLES ONLY (ex labor and kitchen) — this is what feeds Inputs!B19", 6)
put(ws, "A83", "Materials + consumables per unit", BOLD)
put(ws, "D83", "=($D$70+$D$72+$D$73+$D$74+$D$75)*(1+$B$66)", BOLD, USD)
put(ws, "F83", "ingredients + electricity + packaging + delivery + overhead, waste-uplifted. "
    "Time is charged hourly in Scaling instead.", SMALL)

# ============================================================ SENSITIVITY
ws = sheet("Sensitivity", [24, 14, 14, 14, 14, 14, 14])
title(ws, "SENSITIVITY — annual profit at 5 items",
      "Rows = units/day per item. Columns = wholesale % of MSRP. Uses the bake frequency, setup "
      "treatment and cost basis set on Inputs.")
put(ws, "A3", '=" At "&TEXT(Inputs!$B$32,"0")&" bake sessions/week, setup "&Inputs!$B$31&'
              '", materials $"&TEXT(Inputs!$B$20,"0.00")&"/unit, time $"&TEXT(Inputs!$B$30,"0")&"/hr"', BOLD)
hdr(ws, 5, ["Units/day per item", "40%", "45%", "50%", "55%", "60%"])
for j, p in enumerate([0.40, 0.45, 0.50, 0.55, 0.60]):
    put(ws, f"{get_column_letter(2+j)}4", p, BLUE, PCT)
for i, u in enumerate([5, 7, 10, 15, 20, 30, 50]):
    r = 6 + i
    put(ws, f"A{r}", u, BLUE, NUM2)
    ups = f"Inputs!$B$22*$A{r}*7/Inputs!$B$32"
    bps = f"ROUNDUP({ups}/Inputs!$B$24,0)"
    for j in range(5):
        col = get_column_letter(2 + j)
        f = (f"=(Inputs!$B$22*$A{r}*7*SUM(Items!$C$5:$C$9)*{col}$4"
             f"-Inputs!$B$22*$A{r}*7*5*Inputs!$B$20"
             f'-IF(Inputs!$B$21="Yes",IF(Inputs!$B$31="Shared",'
             f"Inputs!$B$32*(Inputs!$B$25+5*{bps}*(Inputs!$B$26+Inputs!$B$27))/60,"
             f"Inputs!$B$32*5*(Inputs!$B$25+{bps}*(Inputs!$B$26+Inputs!$B$27))/60)"
             f"*Inputs!$B$30,0))*52")
        put(ws, f"{col}{r}", f, BLACK, USD0)
        ws[f"{col}{r}"].border = THIN
put(ws, "A14", "Any cell at or below zero means five SKUs at that volume and that wholesale rate "
               "do not cover cost.", SMALL)
put(ws, "A15", "If the grid is all red, the fix is not price — it is Inputs!B32 (bake less often, "
               "bake bigger) or Inputs!B24 (bigger batches).", SMALL)

# ============================================================ BREAKEVEN
ws = sheet("BreakEven", [40, 15, 15, 15, 15, 15, 44])
title(ws, "BREAK-EVEN — can wholesale work at all?",
      "Time cost per unit has a floor: even with infinitely many units, you still pay prep+bake "
      "for every 20-unit batch. If that floor exceeds your margin, no volume rescues it.")

band(ws, 4, "A · THE FLOOR   (what a unit costs when volume is unlimited and setup rounds to nothing)", 7)
rows = [
    ("Blended MSRP across the 5 items", "=AVERAGE(Items!$C$5:$C$9)", USD, "Items tab"),
    ("Blended wholesale price", "=$B$5*Inputs!$B$14", USD, "at Inputs!B14"),
    ("Materials per unit (basis in use)", "=Inputs!$B$20", USD, "Flat or Computed, per Inputs!B18"),
    ("Margin available for time, per unit", "=$B$6-$B$7", USD, "everything left to pay labor+kitchen"),
    ("Minutes per unit at a FULL batch", "=(Inputs!$B$26+Inputs!$B$27)/Inputs!$B$24", NUM2,
     "(prep+bake)/yield — setup excluded, this is the floor"),
    ("Floor time cost per unit", "=$B$9/60*Inputs!$B$30", USD, "at the combined hourly rate"),
    ("Contribution per unit at the floor", "=$B$8-$B$10", USD, "POSITIVE = viable at high volume"),
]
for i, (l, f, fmt, note) in enumerate(rows):
    r = 5 + i
    put(ws, f"A{r}", l, BOLD if i in (3, 6) else BLACK)
    put(ws, f"B{r}", f, BOLD if i in (3, 6) else BLACK, fmt)
    put(ws, f"G{r}", note, SMALL)
put(ws, "A12", '=IF($B$11>0,"VIABLE at high volume — the question is only how fast you get there.",'
               '"NOT VIABLE AT ANY VOLUME with these settings. Fix one of: batch yield, hourly rate, '
               'materials cost, or wholesale %.")', BOLD)

band(ws, 14, "B · WHAT WOULD HAVE TO CHANGE", 7)
put(ws, "A15", "Batch yield needed to break even", BOLD)
put(ws, "B15", "=IFERROR((Inputs!$B$26+Inputs!$B$27)/60*Inputs!$B$30/$B$8,0)", BOLD, NUM2)
put(ws, "G15", "units per batch, at the current rate and margin", SMALL)
put(ws, "A16", "Combined $/hr the current yield can carry", BOLD)
put(ws, "B16", "=IFERROR($B$8/($B$9/60),0)", BOLD, USD)
put(ws, "G16", "vs your $80/hr — the gap is the problem", SMALL)
put(ws, "A17", "Wholesale % needed at the floor", BOLD)
put(ws, "B17", "=IFERROR(($B$7+$B$10)/$B$5,0)", BOLD, PCT)
put(ws, "G17", "share of MSRP that just covers materials + floor time", SMALL)
put(ws, "A18", "Materials budget at 50% wholesale", BOLD)
put(ws, "B18", "=$B$6-$B$10", BOLD, USD)
put(ws, "G18", "what a unit's ingredients must cost to break even at the floor", SMALL)

band(ws, 20, "C · COST PER UNIT BY SESSION SIZE   (one SKU, one bake session)", 7)
hdr(ws, 21, ["Units per session", "Batches", "Minutes", "Time $/unit", "All-in $/unit",
             "Contribution $/unit", "Profit per session"])
for i, u in enumerate([10, 20, 40, 60, 100, 140, 200, 300, 400]):
    r = 22 + i
    put(ws, f"A{r}", u, BLUE, NUM2)
    put(ws, f"B{r}", f"=ROUNDUP($A{r}/Inputs!$B$24,0)", BLACK, NUM2)
    put(ws, f"C{r}", f"=Inputs!$B$25+$B{r}*(Inputs!$B$26+Inputs!$B$27)", BLACK, NUM2)
    put(ws, f"D{r}", f"=IFERROR($C{r}/60*Inputs!$B$30/$A{r},0)", BLACK, USD)
    put(ws, f"E{r}", f"=$D{r}+Inputs!$B$20", BLACK, USD)
    put(ws, f"F{r}", f"=$B$6-$E{r}", BLACK, USD)
    put(ws, f"G{r}", f"=$F{r}*$A{r}", BLACK, USD0)
    ws[f"A{r}"].border = THIN
put(ws, "A32", "Column F turns positive at the volume where wholesale finally covers cost. If it "
               "never does, section A already told you why.", SMALL)

band(ws, 34, "D · THE SAME TABLE IF YOU RAISE BATCH YIELD", 7)
put(ws, "A35", "Bigger trays or a bigger oven is the single highest-leverage change: it divides "
               "the 29 min of prep+bake across more units.", SMALL)
hdr(ws, 36, ["Batch yield", "Min/unit at floor", "Floor time $/unit", "Contribution at floor", "", "", ""])
for i, y in enumerate([20, 30, 40, 60, 80, 100]):
    r = 37 + i
    put(ws, f"A{r}", y, BLUE, NUM2)
    put(ws, f"B{r}", f"=(Inputs!$B$26+Inputs!$B$27)/$A{r}", BLACK, NUM2)
    put(ws, f"C{r}", f"=$B{r}/60*Inputs!$B$30", BLACK, USD)
    put(ws, f"D{r}", f"=$B$8-$C{r}", BLACK, USD)
    ws[f"A{r}"].border = THIN


# ============================================================ PATH
ws = sheet("Path", [30, 13, 13, 13, 13, 13, 13, 40])
title(ws, "THE PATH FROM 1 SKU TO 5",
      "You have one product in one account today. This grid separates the two ways to grow: "
      "more SKUs on one shelf (down) versus the same SKUs in more accounts (across).")
put(ws, "A3", '=" Annual profit · "&TEXT(Inputs!$B$12,"0")&" units/day per item per account · "'
              '&TEXT(Inputs!$B$32,"0")&" bakes/week · setup "&Inputs!$B$31&" · batch yield "'
              '&TEXT(Inputs!$B$24,"0")&" · materials $"&TEXT(Inputs!$B$20,"0.00")&" · time $"'
              '&TEXT(Inputs!$B$30,"0")&"/hr"', BOLD)

ACCS = [1, 2, 3, 5, 8, 12]
band(ws, 5, "ANNUAL PROFIT   rows = SKUs carried · columns = wholesale accounts", 8)
hdr(ws, 6, ["SKUs"] + [f"{a} account" + ("s" if a > 1 else "") for a in ACCS] + [""])
for i, a in enumerate(ACCS):
    put(ws, f"{get_column_letter(2+i)}5", a, BLUE, NUM2)
for n in range(1, 6):
    r = 6 + n
    put(ws, f"A{r}", f"{n} SKU" + ("s" if n > 1 else ""), BOLD)
    for j, a in enumerate(ACCS):
        col = get_column_letter(2 + j)
        ups = f"{col}$5*Inputs!$B$12*7/Inputs!$B$32"
        bps = f"ROUNDUP({ups}/Inputs!$B$24,0)"
        f = (f"=({col}$5*Inputs!$B$12*7*INDEX(Items!$I$5:$I$9,{n})"
             f"-{col}$5*Inputs!$B$12*7*INDEX(Items!$J$5:$J$9,{n})"
             f'-IF(Inputs!$B$21="Yes",IF(Inputs!$B$31="Shared",'
             f"Inputs!$B$32*(Inputs!$B$25+{n}*{bps}*(Inputs!$B$26+Inputs!$B$27))/60,"
             f"Inputs!$B$32*{n}*(Inputs!$B$25+{bps}*(Inputs!$B$26+Inputs!$B$27))/60)"
             f"*Inputs!$B$30,0))*52")
        put(ws, f"{col}{r}", f, BLACK, USD0)
        ws[f"{col}{r}"].border = THIN
put(ws, "H7", "Read across, not down: accounts fill batches, SKUs fill sessions.", SMALL)

band(ws, 13, "UNITS PER BAKE SESSION  (the number that decides everything)", 8)
hdr(ws, 14, ["SKUs"] + [f"{a} acct" + ("s" if a > 1 else "") for a in ACCS] + [""])
for n in range(1, 6):
    r = 14 + n
    put(ws, f"A{r}", f"{n} SKU" + ("s" if n > 1 else ""), BOLD)
    for j, a in enumerate(ACCS):
        col = get_column_letter(2 + j)
        put(ws, f"{col}{r}", f"={n}*{col}$5*Inputs!$B$12*7/Inputs!$B$32", BLACK, NUM2)
        ws[f"{col}{r}"].border = THIN
put(ws, "H15", '=" A full batch is "&TEXT(Inputs!$B$24,"0")&" units. Anything not a clean multiple '
               'wastes a partial batch\'s prep and bake time."', SMALL)

band(ws, 21, "WHAT HERE & THERE ACTUALLY SELLS   (measured — your competition on that shelf)", 8)
hdr(ws, 22, ["Benchmark", "Value", "", "", "", "", "", "Source"])
bench = [
    ("Your madeleine, units/day at LM", 8.7, "Store Sales tab, Aug 1-7 average", NUM2),
    ("Total pastry units/day at LM (est)", 84, "GFS invoice 52 units / 62% GFS share of the mix", NUM2),
    ("GFS SKUs on that invoice", 14, "invoice 1066099707, 07-19-26", NUM2),
    ("Average GFS SKU, units/day", 3.7, "52 units across 14 SKUs", NUM2),
    ("Best GFS SKU (ham & gruyere)", 10, "the ceiling for one SKU at one cafe", NUM2),
    ("GFS wholesale price range", 4.35, "$2.30 cookie to $4.35 ham & gruyere - your price band", USD),
    ("Your wholesale at 50% of $5", 2.50, "sits inside that band", USD),
]
for i, (l, v, src, fmt) in enumerate(bench):
    r = 23 + i
    put(ws, f"A{r}", l, BLACK)
    put(ws, f"B{r}", v, BLUE, fmt)
    put(ws, f"H{r}", src, SMALL)
    ws[f"A{r}"].border = THIN
put(ws, "A31", "A single cafe absorbs roughly 3-10 units/day of any one pastry SKU. Five SKUs at "
               "one account is therefore ~20-50 units/day total — not enough to fill batches.", SMALL)
put(ws, "A32", "The same five SKUs across five accounts is 100-250 units/day on the SAME bake "
               "sessions. That is where the money is.", SMALL)



# ============================================================ CAPACITY
ws = sheet("Capacity", [24, 11, 11, 11, 11, 13, 13, 13, 13, 13, 12, 13])
title(ws, "CAPACITY — what a filled week is worth",
      "Drive the model from HOURS instead of volume. You buy kitchen time by the hour, so the "
      "real question is what each bake hour produces and whether it clears $80.")
put(ws, "A3", '=" "&TEXT(Inputs!$B$32,"0")&" sessions/week · batch yield "&TEXT(Inputs!$B$24,"0")&'
              '" · "&TEXT(Inputs!$B$26+Inputs!$B$27,"0")&" min per batch · materials $"&'
              'TEXT(Inputs!$B$20,"0.00")&"/unit · time $"&TEXT(Inputs!$B$30,"0")&"/hr"', BOLD)

band(ws, 5, "A · THE HOURLY MATH   (this one block explains every number below)", 12)
hm = [
    ("Minutes per batch (prep + bake)", "=Inputs!$B$26+Inputs!$B$27", NUM2, "PREP_MIN + BAKE_MIN"),
    ("Units per batch", "=Inputs!$B$24", NUM2, "BATCH_YIELD"),
    ("Units produced per bake hour", "=60/$B$6*$B$7", NUM2, "setup excluded — the theoretical rate"),
    ("Blended wholesale price", "=AVERAGE(Items!$C$5:$C$9)*Inputs!$B$14", USD, "Items tab x Inputs!B14"),
    ("Materials per unit", "=Inputs!$B$20", USD, "Flat or Computed per Inputs!B18"),
    ("Contribution per unit", "=$B$9-$B$10", USD, "before any time cost"),
    ("Gross contribution per bake hour", "=$B$8*$B$11", USD, "what an hour of baking earns"),
    ("Cost of a bake hour", "=Inputs!$B$30", USD, "labor + commercial kitchen"),
    ("NET PROFIT PER BAKE HOUR", "=$B$12-$B$13", USD, "if this is negative, no schedule fixes it"),
]
for i, (l, f, fmt, note) in enumerate(hm):
    r = 6 + i
    put(ws, f"A{r}", l, BOLD if i == 8 else BLACK)
    put(ws, f"B{r}", f, BOLD if i == 8 else BLACK, fmt)
    put(ws, f"D{r}", note, SMALL)
put(ws, "A15", '=IF($B$14>0,"Every additional bake hour adds $"&TEXT($B$14,"0.00")&" of profit. '
               'Fill the schedule.","A bake hour LOSES $"&TEXT(-$B$14,"0.00")&". Filling the '
               'schedule multiplies the loss — fix yield, rate, or price first.")', BOLD)

band(ws, 17, "B · WEEKLY SCHEDULE SCENARIOS", 12)
put(ws, "A18", "SKUs carried", BOLD)
put(ws, "B18", 5, BLUE, NUM2, YFILL)
put(ws, "C18", "used only for the accounts-needed column", SMALL)
hdr(ws, 19, ["Schedule", "Hours/week", "Batches/week", "Units/week", "Units/day", "Revenue/week",
             "Materials/week", "Time cost/week", "Profit/week", "Profit/year", "$/bake hour",
             "Accounts needed"])
HRS = [10, 20, 30, 40, 50, 60, 80, 100]
for i, h in enumerate(HRS):
    r = 20 + i
    lbl = f"{h} hours" + ("  (one full-time baker)" if h == 40 else
                          "  (two bakers)" if h == 80 else "")
    put(ws, f"A{r}", lbl, BOLD)
    put(ws, f"B{r}", h, BLUE, NUM2)
    put(ws, f"C{r}", f"=MAX(0,ROUNDDOWN(($B{r}*60-Inputs!$B$32*Inputs!$B$25)/$B$6,0))", BLACK, NUM2)
    put(ws, f"D{r}", f"=$C{r}*Inputs!$B$24", BLACK, NUM2)
    put(ws, f"E{r}", f"=$D{r}/7", BLACK, NUM2)
    put(ws, f"F{r}", f"=$D{r}*$B$9", BLACK, USD)
    put(ws, f"G{r}", f"=$D{r}*$B$10", BLACK, USD)
    put(ws, f"H{r}", f'=IF(Inputs!$B$21="Yes",$B{r}*Inputs!$B$30,0)', BLACK, USD)
    put(ws, f"I{r}", f"=$F{r}-$G{r}-$H{r}", BLACK, USD)
    put(ws, f"J{r}", f"=$I{r}*52", BLACK, USD0)
    put(ws, f"K{r}", f"=IFERROR($I{r}/$B{r},0)", BLACK, USD)
    put(ws, f"L{r}", f"=IFERROR($E{r}/($B$18*Inputs!$B$12),0)", BLACK, NUM2)
    for col in "ABCDEFGHIJKL":
        ws[f"{col}{r}"].border = THIN
put(ws, "A29", '="Accounts needed = units/day divided by ("&TEXT($B$18,"0")&" SKUs x "&'
               'TEXT(Inputs!$B$12,"0")&" units/day per SKU per account). Production is worthless '
               'if nobody is buying it."', SMALL)
put(ws, "A30", "Past 40 hrs/week you are buying a second baker and probably a second oven. The "
               "model just keeps charging $/hr — it does not know you are tired.", SMALL)
put(ws, "A31", "Setup is charged once per session per week, not per hour, so longer sessions "
               "amortise it away. That is why $/bake hour improves with hours.", SMALL)

band(ws, 33, "C · THE SAME SCHEDULE AT DIFFERENT BATCH YIELDS   (profit/year)", 12)
hdr(ws, 34, ["Hours/week", "Yield 20", "Yield 30", "Yield 40", "Yield 60", "Yield 80", "", "", "", "", "", ""])
for j, y in enumerate([20, 30, 40, 60, 80]):
    put(ws, f"{get_column_letter(2+j)}33", y, BLUE, NUM2)
for i, h in enumerate(HRS):
    r = 35 + i
    put(ws, f"A{r}", h, BLUE, NUM2)
    for j in range(5):
        col = get_column_letter(2 + j)
        batches = f"MAX(0,ROUNDDOWN(($A{r}*60-Inputs!$B$32*Inputs!$B$25)/$B$6,0))"
        units = f"{batches}*{col}$33"
        f = (f"=({units}*$B$9-{units}*$B$10"
             f'-IF(Inputs!$B$21="Yes",$A{r}*Inputs!$B$30,0))*52')
        put(ws, f"{col}{r}", f, BLACK, USD0)
        ws[f"{col}{r}"].border = THIN
put(ws, "A44", "Batch yield is the whole business. At 20 per batch you produce ~41 units an hour; "
               "at 40 you produce ~83 for the same $80.", SMALL)


wb.move_sheet("Inputs", offset=-10)
wb.save("/Users/andrew/kairos-wholesale/wholesale_model.xlsx")
print("written")
